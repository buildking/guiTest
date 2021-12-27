from openpyxl.styles import PatternFill
import openpyxl
from tkinter import *
import pandas as pd
import wemade_log
import datetime
import time as Time
from os import *
from wemade_main import *
import numpy as np

logger = wemade_log.setLogging("compare")

#end, new excel file compare
def xlCompare(_text=None):

    time = str(datetime.datetime.now())[0:-7]
    _text.insert(END, f"\n[{time}] 엑셀파일을 불러옵니다.\n")

    #폴더 안 파일의 리스트 생성
    excelList = listdir('input')
    logger.info(excelList)
    print(excelList)
    time = str(datetime.datetime.now())[0:-7]
    _text.insert(END, f"[{time}] 파일 리스트\n{excelList}\n")

    #폴더 안에 파일이 2개가 아니라면 에러메세지 띄우기
    if len(excelList) != 2:
        shortError('input')
        return

    # #리스트 안의 파일 하나씩 꺼내기
    # #첫번째 파일
    # try:
    #     endFrame = pd.DataFrame()
    #     #엑셀 모든 시트를 합쳐서 받는다.
    #     endFrame_dic = pd.read_excel(f'./input/{excelList[0]}', sheet_name=None, dtype=str)
    #     endFrame = pd.DataFrame.from_dict([endFrame_dic])
    #     print(endFrame)
    #     print(type(endFrame))
    #     logger.info(endFrame)
    #     time = str(datetime.datetime.now())[0:-7]
    #     _text.insert(END, f"[{time}] {endFrame}\n")
    # except:
    #     inputError()
    # #두번째 파일
    # try:
    #     newFrame = pd.DataFrame()
    #     #엑셀 모든 시트를 합쳐서 받는다.
    #     newFrame_dic = pd.read_excel(f'./input/{excelList[1]}', sheet_name=None, dtype=str)
    #     newFrame = pd.DataFrame.from_dict([newFrame_dic])
    #     print(newFrame)
    #     print(type(newFrame))
    #     logger.info(newFrame)
    #     time = str(datetime.datetime.now())[0:-7]
    #     _text.insert(END, f"[{time}] {newFrame}\n")
    # except:
    #     inputError()
    #
    # resultFrame = newFrame.copy()

    xl = pd.ExcelFile(f'./input/{excelList[1]}')
    #엑셀 sheet 개수
    xlSheet = len(xl.sheet_names)
    print(xlSheet)
    #엑셀 sheet별 column*row
    xlShape = []


    # 데이터가 다른 cell의 column, row 저장(append로 추가)
    coreRowNum = []
    coreRowNum.clear()
    coreRowNumPrint = []
    coreRowNumPrint.clear()

    for sheetNum in range(xlSheet):
        firExcel = pd.read_excel(f'./input/{excelList[0]}', sheet_name = sheetNum, dtype = str)
        secExcel = pd.read_excel(f'./input/{excelList[1]}', sheet_name = sheetNum, dtype = str)
        print(secExcel)

        #딕셔너리화
        f_list = firExcel.values.tolist()
        s_list = secExcel.values.tolist()

        print('*'*100)
        print(s_list)

        #sheet별 column, row
        xlRow = len(secExcel.index)
        xlColumn = len(secExcel.columns)

        xlShape.append([xlRow, xlColumn])
        print(xlRow, xlColumn)


        i = 0
        j = 0
        for _ in range(xlRow):
            j = 0
            for _ in range(xlColumn):
                ######비교 로직 작성

                if str(f_list[i][j]) != str(s_list[i][j]):
                    coreRowNum.append([sheetNum, i, j])
                    coreRowNumPrint.append([sheetNum, i + 1, j + 1])
                    print('!!!')
                    print(sheetNum,i,j)

                ######
                j = j + 1
            i = i + 1



    time = str(datetime.datetime.now())[0:-7]
    _text.insert(END, f"\n[{time}] 데이터 변경점 탐색을 시작합니다.\n")

    start = Time.time()

    #
    # #데이터가 다른 end data와 new data의 column, row 저장(append로 추가)
    # coreRowNum = []
    # coreRowNum.clear()
    # coreRowNumPrint = []
    # coreRowNumPrint.clear()
    #
    # print(type(newFrame))
    #
    # columnList = newFrame.columns.tolist()
    # #dataframe을 dictionary type으로 변환
    # b_dict = endFrame.to_dict("list")
    # a_dict = newFrame.to_dict("list")
    # rowList = len(a_dict)
    # i = 0
    # j = 0
    # for comCol in columnList:
    #     j = 0
    #     bList = b_dict[f'{comCol}']
    #     aList = a_dict[f'{comCol}']
    #     for comRow in range(rowList):
    #         ######비교 로직 작성
    #         if bList[comRow] is not aList[comRow]:
    #             coreRowNum.append([i, j])
    #             coreRowNumPrint.append([i+1, j+1])
    #         ######
    #         j = j + 1
    #     i = i + 1

    end = Time.time()
    logger.info("excute time >>>>> " + str(end - start) + "s")

    logger.debug("변경된 데이터 리스트")
    logger.debug(coreRowNumPrint)
    logger.debug(len(coreRowNumPrint))
    print("변경된 데이터 리스트")
    print(coreRowNumPrint)
    print(len(coreRowNumPrint))

    #데이터 이격 생긴 부분 컬럼, 로우만 갈무리 해서 엑셀 생성
    coreRowNumExcel = pd.DataFrame(coreRowNumPrint, columns=['sheet','row','column'])
    coreRowNumExcel.to_excel('./output/excel_tryout_rows.xlsx')

    #제목에 넣을 날짜
    finishStamp = str(datetime.datetime.now())[0:10]
    #cell에 색칠할 엑셀 output에 저장
    xlTemp = openpyxl.load_workbook(filename = f'./input/{excelList[1]}')
    xlTemp.save(filename = f'./output/{excelList[0],excelList[1],finishStamp}.xlsx')
    xlTemp.close()


    #cell 색칠 시작
    time = str(datetime.datetime.now())[0:-7]
    _text.insert(END, f"[{time}] painting start")
    logger.info("painting start")
    print("Excel file painting start")

    #노란색 선언
    y_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')

    #openpyxl을 통해 엑셀시트를 연다.
    pyxl = openpyxl.load_workbook(filename=f'./output/{excelList[0],excelList[1],finishStamp}.xlsx')

    for cellSheet, cellRow, cellCol in coreRowNumPrint:

        pyxlSheet = pyxl.worksheets[cellSheet]

        #i=column, j=row
        pyxlSheet.cell(cellRow+1, cellCol).fill = y_color

    pyxl.save(filename=f'./output/{excelList[0],excelList[1],finishStamp}.xlsx')

    pyxl.close()


    time = str(datetime.datetime.now())[0:-7]
    _text.insert(END, f"[{time}] painting complete")
    logger.info("painting complete")
    print("Excel file painting complete")

    time = str(datetime.datetime.now())[0:-7]
    _text.insert(END, f"[{time}] 데이터 변경점 검색을 종료합니다.\n")
    _text.see(END)